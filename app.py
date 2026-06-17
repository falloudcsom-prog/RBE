from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file, abort
from flask_sqlalchemy import SQLAlchemy
from flask_bcrypt import Bcrypt
from functools import wraps
import os
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import uuid

app = Flask(__name__)
app.config['SECRET_KEY'] = 'rne_sn_2024_xK9#mPqL!vR7'
# Chemin absolu pour la base SQLite — fonctionne quel que soit le répertoire courant
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'instance', 'declarations.db')
os.makedirs(os.path.join(BASE_DIR, 'instance'), exist_ok=True)
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_PATH}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024

db = SQLAlchemy(app)
bcrypt = Bcrypt(app)

ALLOWED_EXTENSIONS = {'pdf', 'zip', 'docx', 'xlsx'}


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), nullable=False)
    prenom = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), default='utilisateur')
    actif = db.Column(db.Boolean, default=True)
    date_creation = db.Column(db.DateTime, default=datetime.utcnow)
    declarations = db.relationship('Declaration', backref='auteur', lazy=True)


class Declaration(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    reference = db.Column(db.String(50), unique=True, nullable=False)
    date_saisie = db.Column(db.DateTime, default=datetime.utcnow)
    utilisateur_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    statut = db.Column(db.String(20), default='En attente')

    denomination_sociale = db.Column(db.String(200))
    ninea = db.Column(db.String(50))
    rccm = db.Column(db.String(100))
    forme_juridique = db.Column(db.String(100))
    pays = db.Column(db.String(100))
    ville = db.Column(db.String(100))
    adresse = db.Column(db.String(300))
    telephone = db.Column(db.String(30))
    email_entreprise = db.Column(db.String(120))
    cotee_bourse = db.Column(db.String(20))
    filiale_cotee = db.Column(db.String(20))
    entreprise_etat = db.Column(db.String(20))
    autre_entite = db.Column(db.String(20))
    nom_maison_mere = db.Column(db.String(200))
    place_boursiere = db.Column(db.String(200))
    numero_isin = db.Column(db.String(100))
    pourcentage_participation_mere = db.Column(db.String(20))
    region = db.Column(db.String(100))

    nom_beneficiaire = db.Column(db.String(200))
    prenom_beneficiaire = db.Column(db.String(200))
    sexe = db.Column(db.String(20))
    date_naissance = db.Column(db.String(20))
    lieu_naissance = db.Column(db.String(100))
    nationalite = db.Column(db.String(100))
    pays_residence = db.Column(db.String(100))
    adresse_personnelle = db.Column(db.String(300))
    adresse_professionnelle = db.Column(db.String(300))
    numero_cni = db.Column(db.String(50))
    numero_passeport = db.Column(db.String(50))
    ppe = db.Column(db.String(20))
    relation_ppe = db.Column(db.String(20))
    fonction_ppe = db.Column(db.String(200))
    categorie_ppe = db.Column(db.String(100))
    date_debut_fonction = db.Column(db.String(20))
    date_fin_fonction = db.Column(db.String(20))
    nature_relation_ppe = db.Column(db.Text)

    parts_directes = db.Column(db.String(20))
    nombre_parts_directes = db.Column(db.String(50))
    pourcentage_participation = db.Column(db.String(20))
    droits_vote_directs = db.Column(db.String(20))
    nombre_voix = db.Column(db.String(50))
    pourcentage_voix = db.Column(db.String(20))
    parts_indirectes = db.Column(db.String(20))
    nombre_parts_indirectes = db.Column(db.String(50))
    pourcentage_participation_indirecte = db.Column(db.String(20))
    droits_vote_indirects = db.Column(db.String(20))
    representant_legal = db.Column(db.String(20))
    date_beneficiaire = db.Column(db.String(20))
    autres_beneficiaires = db.Column(db.String(20))
    intermediaires = db.Column(db.Text)

    fichiers = db.Column(db.Text)
    lien_drive = db.Column(db.String(500))
    lien_zip = db.Column(db.String(500))

    certifie_par = db.Column(db.String(200))
    fonction_certifiant = db.Column(db.String(200))


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Veuillez vous connecter.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        user = User.query.get(session['user_id'])
        if not user or user.role != 'admin':
            flash('Accès réservé aux administrateurs.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def generate_reference():
    now = datetime.now()
    return f"RNE-{now.year}-{now.strftime('%m%d')}-{str(uuid.uuid4())[:6].upper()}"


def get_stats(user_id=None):
    """Statistiques globales (admin) ou filtrées par utilisateur."""
    q = Declaration.query
    if user_id:
        q = q.filter_by(utilisateur_id=user_id)

    total = q.count()

    q_today = q.filter(db.func.date(Declaration.date_saisie) == datetime.today().date())
    today = q_today.count()

    if user_id:
        entreprises  = db.session.query(Declaration.ninea).filter_by(utilisateur_id=user_id).distinct().count()
        beneficiaires = db.session.query(Declaration.nom_beneficiaire).filter_by(utilisateur_id=user_id).distinct().count()
        ppe_count    = Declaration.query.filter_by(utilisateur_id=user_id).filter(Declaration.ppe == 'Oui').count()
    else:
        entreprises  = db.session.query(Declaration.ninea).distinct().count()
        beneficiaires = db.session.query(Declaration.nom_beneficiaire).distinct().count()
        ppe_count    = Declaration.query.filter(Declaration.ppe == 'Oui').count()

    return {
        'total': total,
        'today': today,
        'entreprises': entreprises,
        'beneficiaires': beneficiaires,
        'ppe': ppe_count
    }


@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        user = User.query.filter_by(email=email).first()
        if user and bcrypt.check_password_hash(user.password, password) and user.actif:
            session['user_id'] = user.id
            session['user_role'] = user.role
            session['user_nom'] = f"{user.prenom} {user.nom}"
            return redirect(url_for('dashboard'))
        flash('Identifiants incorrects ou compte désactivé.', 'danger')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    user = User.query.get(session['user_id'])
    est_admin = (user.role == 'admin')

    # Statistiques filtrées selon le rôle
    uid_filtre = None if est_admin else user.id
    stats = get_stats(user_id=uid_filtre)

    # Déclarations récentes filtrées
    if est_admin:
        declarations = Declaration.query.order_by(Declaration.date_saisie.desc()).limit(10).all()
    else:
        declarations = Declaration.query.filter_by(utilisateur_id=user.id).order_by(Declaration.date_saisie.desc()).limit(10).all()

    from sqlalchemy import extract, func

    # Graphique par mois — filtré par utilisateur si nécessaire
    q_mois = db.session.query(
        extract('month', Declaration.date_saisie).label('mois'),
        func.count(Declaration.id).label('count')
    )
    if not est_admin:
        q_mois = q_mois.filter(Declaration.utilisateur_id == user.id)
    mois_data = q_mois.group_by('mois').all()

    mois_labels = ['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Jun', 'Jul', 'Aoû', 'Sep', 'Oct', 'Nov', 'Déc']
    mois_counts = [0] * 12
    for m, c in mois_data:
        if m:
            mois_counts[int(m) - 1] = c

    # Graphique formes juridiques — filtré
    q_formes = db.session.query(
        Declaration.forme_juridique, func.count(Declaration.id)
    )
    if not est_admin:
        q_formes = q_formes.filter(Declaration.utilisateur_id == user.id)
    formes_data = q_formes.group_by(Declaration.forme_juridique).all()

    return render_template('dashboard.html',
        stats=stats,
        declarations=declarations,
        mois_labels=json.dumps(mois_labels),
        mois_counts=json.dumps(mois_counts),
        formes_data=json.dumps([{'label': f or 'Non renseigné', 'value': c} for f, c in formes_data]),
        user=user,
        est_admin=est_admin
    )


@app.route('/declaration/nouvelle', methods=['GET', 'POST'])
@login_required
def nouvelle_declaration():
    if request.method == 'POST':
        try:
            user_id = session['user_id']
            d = Declaration(
                reference=generate_reference(),
                utilisateur_id=user_id,
                denomination_sociale=request.form.get('denomination_sociale'),
                ninea=request.form.get('ninea'),
                rccm=request.form.get('rccm'),
                forme_juridique=request.form.get('forme_juridique'),
                pays=request.form.get('pays'),
                ville=request.form.get('ville'),
                adresse=request.form.get('adresse'),
                telephone=request.form.get('telephone'),
                email_entreprise=request.form.get('email_entreprise'),
                cotee_bourse=request.form.get('cotee_bourse'),
                filiale_cotee=request.form.get('filiale_cotee'),
                entreprise_etat=request.form.get('entreprise_etat'),
                autre_entite=request.form.get('autre_entite'),
                nom_maison_mere=request.form.get('nom_maison_mere'),
                place_boursiere=request.form.get('place_boursiere'),
                numero_isin=request.form.get('numero_isin'),
                pourcentage_participation_mere=request.form.get('pourcentage_participation_mere'),
                region=request.form.get('region'),
                nom_beneficiaire=request.form.get('nom_beneficiaire'),
                prenom_beneficiaire=request.form.get('prenom_beneficiaire'),
                sexe=request.form.get('sexe'),
                date_naissance=request.form.get('date_naissance'),
                lieu_naissance=request.form.get('lieu_naissance'),
                nationalite=request.form.get('nationalite'),
                pays_residence=request.form.get('pays_residence'),
                adresse_personnelle=request.form.get('adresse_personnelle'),
                adresse_professionnelle=request.form.get('adresse_professionnelle'),
                numero_cni=request.form.get('numero_cni'),
                numero_passeport=request.form.get('numero_passeport'),
                ppe=request.form.get('ppe'),
                relation_ppe=request.form.get('relation_ppe'),
                fonction_ppe=request.form.get('fonction_ppe'),
                categorie_ppe=request.form.get('categorie_ppe'),
                date_debut_fonction=request.form.get('date_debut_fonction'),
                date_fin_fonction=request.form.get('date_fin_fonction'),
                nature_relation_ppe=request.form.get('nature_relation_ppe'),
                parts_directes=request.form.get('parts_directes'),
                nombre_parts_directes=request.form.get('nombre_parts_directes'),
                pourcentage_participation=request.form.get('pourcentage_participation'),
                droits_vote_directs=request.form.get('droits_vote_directs'),
                nombre_voix=request.form.get('nombre_voix'),
                pourcentage_voix=request.form.get('pourcentage_voix'),
                parts_indirectes=request.form.get('parts_indirectes'),
                nombre_parts_indirectes=request.form.get('nombre_parts_indirectes'),
                pourcentage_participation_indirecte=request.form.get('pourcentage_participation_indirecte'),
                droits_vote_indirects=request.form.get('droits_vote_indirects'),
                representant_legal=request.form.get('representant_legal'),
                date_beneficiaire=request.form.get('date_beneficiaire'),
                autres_beneficiaires=request.form.get('autres_beneficiaires'),
                certifie_par=request.form.get('certifie_par'),
                fonction_certifiant=request.form.get('fonction_certifiant'),
                lien_drive=request.form.get('lien_drive'),
            )

            fichiers_list = []
            if 'documents' in request.files:
                files = request.files.getlist('documents')
                ninea_safe = (request.form.get('ninea') or 'inconnu').replace('/', '_').replace(' ', '_')
                upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], f'Entreprise_{ninea_safe}')
                os.makedirs(upload_dir, exist_ok=True)
                for file in files:
                    if file and file.filename and allowed_file(file.filename):
                        ts = datetime.now().strftime('%Y%m%d%H%M%S')
                        fname = f"{ts}_{file.filename}"
                        file.save(os.path.join(upload_dir, fname))
                        fichiers_list.append({'nom': file.filename, 'path': f"Entreprise_{ninea_safe}/{fname}"})
            d.fichiers = json.dumps(fichiers_list)

            db.session.add(d)
            db.session.commit()
            export_to_excel()
            flash('Déclaration enregistrée avec succès.', 'success')
            return redirect(url_for('detail_declaration', id=d.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de l\'enregistrement : {str(e)}', 'danger')

    return render_template('declaration.html')


@app.route('/declaration/<int:id>')
@login_required
def detail_declaration(id):
    d = Declaration.query.get_or_404(id)
    user = User.query.get(session['user_id'])
    if user.role != 'admin' and d.utilisateur_id != user.id:
        flash('Accès non autorisé.', 'danger')
        return redirect(url_for('dashboard'))
    fichiers = json.loads(d.fichiers) if d.fichiers else []
    return render_template('detail.html', d=d, fichiers=fichiers, user=user)


@app.route('/declaration/<int:id>/modifier', methods=['GET', 'POST'])
@login_required
def modifier_declaration(id):
    d = Declaration.query.get_or_404(id)
    user = User.query.get(session['user_id'])
    if user.role != 'admin' and d.utilisateur_id != user.id:
        flash('Accès non autorisé.', 'danger')
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        champs = ['denomination_sociale','ninea','rccm','forme_juridique','pays','ville','adresse',
                  'telephone','email_entreprise','cotee_bourse','filiale_cotee','entreprise_etat',
                  'autre_entite','nom_maison_mere','place_boursiere','numero_isin',
                  'pourcentage_participation_mere','region','nom_beneficiaire','prenom_beneficiaire',
                  'sexe','date_naissance','lieu_naissance','nationalite','pays_residence',
                  'adresse_personnelle','adresse_professionnelle','numero_cni','numero_passeport',
                  'ppe','relation_ppe','fonction_ppe','categorie_ppe','date_debut_fonction',
                  'date_fin_fonction','nature_relation_ppe','parts_directes','nombre_parts_directes',
                  'pourcentage_participation','droits_vote_directs','nombre_voix','pourcentage_voix',
                  'parts_indirectes','nombre_parts_indirectes','pourcentage_participation_indirecte',
                  'droits_vote_indirects','representant_legal','date_beneficiaire','autres_beneficiaires',
                  'certifie_par','fonction_certifiant','lien_drive']
        for champ in champs:
            setattr(d, champ, request.form.get(champ))
        db.session.commit()
        export_to_excel()
        flash('Déclaration modifiée avec succès.', 'success')
        return redirect(url_for('detail_declaration', id=d.id))
    fichiers = json.loads(d.fichiers) if d.fichiers else []
    return render_template('declaration.html', d=d, fichiers=fichiers, edit=True)


@app.route('/declaration/<int:id>/supprimer', methods=['POST'])
@admin_required
def supprimer_declaration(id):
    d = Declaration.query.get_or_404(id)
    db.session.delete(d)
    db.session.commit()
    export_to_excel()
    flash('Déclaration supprimée.', 'success')
    return redirect(url_for('liste_declarations'))


@app.route('/declaration/<int:id>/statut', methods=['POST'])
@admin_required
def changer_statut(id):
    d = Declaration.query.get_or_404(id)
    d.statut = request.form.get('statut', 'En attente')
    db.session.commit()
    return jsonify({'success': True, 'statut': d.statut})


@app.route('/declarations')
@login_required
def liste_declarations():
    user = User.query.get(session['user_id'])
    page = request.args.get('page', 1, type=int)
    search = request.args.get('q', '')
    region = request.args.get('region', '')
    forme = request.args.get('forme', '')
    ppe = request.args.get('ppe', '')
    statut = request.args.get('statut', '')

    query = Declaration.query
    if user.role != 'admin':
        query = query.filter_by(utilisateur_id=user.id)
    if search:
        query = query.filter(
            db.or_(
                Declaration.denomination_sociale.ilike(f'%{search}%'),
                Declaration.ninea.ilike(f'%{search}%'),
                Declaration.rccm.ilike(f'%{search}%'),
                Declaration.nom_beneficiaire.ilike(f'%{search}%'),
            )
        )
    if region:
        query = query.filter(Declaration.region == region)
    if forme:
        query = query.filter(Declaration.forme_juridique == forme)
    if ppe:
        query = query.filter(Declaration.ppe == ppe)
    if statut:
        query = query.filter(Declaration.statut == statut)

    declarations = query.order_by(Declaration.date_saisie.desc()).paginate(page=page, per_page=20)
    regions = db.session.query(Declaration.region).distinct().all()
    formes = db.session.query(Declaration.forme_juridique).distinct().all()
    return render_template('liste.html', declarations=declarations, user=user,
                           regions=regions, formes=formes, search=search,
                           region=region, forme=forme, ppe=ppe, statut=statut)


@app.route('/telecharger')
@login_required
def telecharger():
    user = User.query.get(session['user_id'])
    if user.role == 'admin':
        declarations = Declaration.query.order_by(Declaration.date_saisie.desc()).all()
    else:
        declarations = Declaration.query.filter_by(utilisateur_id=user.id).order_by(Declaration.date_saisie.desc()).all()
    stats = get_stats()
    return render_template('telecharger.html', declarations=declarations, user=user, stats=stats)


@app.route('/fichier/<int:decl_id>/<path:filename>')
@login_required
def voir_fichier(decl_id, filename):
    d = Declaration.query.get_or_404(decl_id)
    user = User.query.get(session['user_id'])
    if user.role != 'admin' and d.utilisateur_id != user.id:
        abort(403)
    fichiers = json.loads(d.fichiers) if d.fichiers else []
    fichier_trouve = None
    for f in fichiers:
        if isinstance(f, dict) and f.get('path') == filename:
            fichier_trouve = f
            break
        elif isinstance(f, str) and f == filename:
            fichier_trouve = {'path': filename, 'nom': filename}
            break
    if not fichier_trouve:
        abort(404)
    full_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(full_path):
        abort(404)
    ext = filename.rsplit('.', 1)[-1].lower()
    mime_types = {
        'pdf': 'application/pdf',
        'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        'zip': 'application/zip',
    }
    mime = mime_types.get(ext, 'application/octet-stream')
    inline = ext == 'pdf'
    disposition = 'inline' if inline else 'attachment'
    return send_file(full_path, mimetype=mime,
                     download_name=fichier_trouve.get('nom', filename),
                     as_attachment=not inline)


@app.route('/admin')
@admin_required
def admin():
    users = User.query.all()
    stats = get_stats()
    return render_template('admin.html', users=users, stats=stats)


@app.route('/admin/utilisateur/nouveau', methods=['POST'])
@admin_required
def nouveau_utilisateur():
    data = request.form
    if User.query.filter_by(email=data['email']).first():
        flash('Cette adresse email est déjà enregistrée.', 'danger')
        return redirect(url_for('admin'))
    hashed = bcrypt.generate_password_hash(data['password']).decode('utf-8')
    user = User(nom=data['nom'], prenom=data['prenom'], email=data['email'],
                password=hashed, role=data.get('role', 'utilisateur'))
    db.session.add(user)
    db.session.commit()
    flash('Compte utilisateur créé.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/utilisateur/<int:id>/toggle', methods=['POST'])
@admin_required
def toggle_utilisateur(id):
    user = User.query.get_or_404(id)
    user.actif = not user.actif
    db.session.commit()
    return jsonify({'actif': user.actif})


@app.route('/admin/utilisateur/<int:id>/supprimer', methods=['POST'])
@admin_required
def supprimer_utilisateur(id):
    user = User.query.get_or_404(id)
    db.session.delete(user)
    db.session.commit()
    flash('Compte supprimé.', 'success')
    return redirect(url_for('admin'))


@app.route('/export/excel')
@login_required
def export_excel():
    path = export_to_excel()
    return send_file(path, as_attachment=True, download_name='registre_beneficiaires_effectifs.xlsx')


@app.route('/export/csv')
@login_required
def export_csv():
    import csv
    import io
    user = User.query.get(session['user_id'])
    if user.role == 'admin':
        declarations = Declaration.query.all()
    else:
        declarations = Declaration.query.filter_by(utilisateur_id=user.id).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Référence', 'Date', 'Entreprise', 'NINEA', 'RCCM', 'Forme Juridique',
                     'Bénéficiaire', 'Nationalité', 'PPE', 'Participation %', 'Statut'])
    for d in declarations:
        writer.writerow([
            d.reference,
            d.date_saisie.strftime('%d/%m/%Y') if d.date_saisie else '',
            d.denomination_sociale, d.ninea, d.rccm, d.forme_juridique,
            f"{d.nom_beneficiaire or ''} {d.prenom_beneficiaire or ''}".strip(),
            d.nationalite, d.ppe, d.pourcentage_participation, d.statut
        ])
    output.seek(0)
    from flask import Response
    return Response(output.getvalue(), mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment;filename=registre_beneficiaires.csv'})


@app.route('/import/excel', methods=['GET', 'POST'])
@admin_required
def import_excel():
    """Page d'import de déclarations en masse via fichier Excel."""
    if request.method == 'GET':
        return render_template('import_excel.html')

    if 'fichier_excel' not in request.files:
        flash('Aucun fichier sélectionné.', 'danger')
        return redirect(url_for('import_excel'))

    fichier = request.files['fichier_excel']
    if not fichier.filename or not fichier.filename.lower().endswith('.xlsx'):
        flash('Le fichier doit être au format Excel (.xlsx).', 'danger')
        return redirect(url_for('import_excel'))

    mode = request.form.get('mode', 'apercu')  # 'apercu' ou 'importer'

    try:
        wb = openpyxl.load_workbook(fichier, data_only=True)
        ws = wb.active

        # Lire les en-têtes (ligne 1)
        headers_row = [str(ws.cell(1, c).value or '').strip() for c in range(1, ws.max_column + 1)]

        # Colonnes attendues et leur correspondance avec les champs du modèle
        MAPPING_COLONNES = {
            'Dénomination Sociale':          'denomination_sociale',
            'Denomination Sociale':          'denomination_sociale',
            'NINEA':                         'ninea',
            'RCCM':                          'rccm',
            'Forme Juridique':               'forme_juridique',
            'Pays':                          'pays',
            'Région':                        'region',
            'Region':                        'region',
            'Ville':                         'ville',
            'Adresse':                       'adresse',
            'Téléphone':                     'telephone',
            'Telephone':                     'telephone',
            'Email Entreprise':              'email_entreprise',
            'Email':                         'email_entreprise',
            'Cotée en bourse':               'cotee_bourse',
            'Cotee en bourse':               'cotee_bourse',
            'Filiale cotée':                 'filiale_cotee',
            'Filiale cotee':                 'filiale_cotee',
            'Entreprise Etat':               'entreprise_etat',
            "Entreprise d'Etat":             'entreprise_etat',
            'Autre entité':                  'autre_entite',
            'Autre entite':                  'autre_entite',
            'Nom Maison Mère':               'nom_maison_mere',
            'Nom Maison Mere':               'nom_maison_mere',
            'Place Boursière':               'place_boursiere',
            'Place Boursiere':               'place_boursiere',
            'Numéro ISIN':                   'numero_isin',
            'Numero ISIN':                   'numero_isin',
            '% Participation Mère':          'pourcentage_participation_mere',
            '% Participation Mere':          'pourcentage_participation_mere',
            'Nom Bénéficiaire':              'nom_beneficiaire',
            'Nom Beneficiaire':              'nom_beneficiaire',
            'Prénom Bénéficiaire':           'prenom_beneficiaire',
            'Prenom Beneficiaire':           'prenom_beneficiaire',
            'Sexe':                          'sexe',
            'Date Naissance':                'date_naissance',
            'Lieu Naissance':                'lieu_naissance',
            'Nationalité':                   'nationalite',
            'Nationalite':                   'nationalite',
            'Pays Résidence':                'pays_residence',
            'Pays Residence':                'pays_residence',
            'Adresse Personnelle':           'adresse_personnelle',
            'Adresse Professionnelle':       'adresse_professionnelle',
            'N° CNI / Passeport':            'numero_cni',
            'N° CNI':                        'numero_cni',
            'Numero CNI':                    'numero_cni',
            'N° Passeport':                  'numero_passeport',
            'PPE':                           'ppe',
            'Relation PPE':                  'relation_ppe',
            'Fonction PPE':                  'fonction_ppe',
            'Catégorie PPE':                 'categorie_ppe',
            'Categorie PPE':                 'categorie_ppe',
            'Date Début Fonction':           'date_debut_fonction',
            'Date Debut Fonction':           'date_debut_fonction',
            'Date Fin Fonction':             'date_fin_fonction',
            'Parts Directes':                'parts_directes',
            'Nombre Parts Directes':         'nombre_parts_directes',
            '% Participation':               'pourcentage_participation',
            'Participation %':               'pourcentage_participation',
            'Pourcentage Participation':     'pourcentage_participation',
            'Droits Vote Directs':           'droits_vote_directs',
            'Nombre Voix':                   'nombre_voix',
            '% Voix':                        'pourcentage_voix',
            'Pourcentage Voix':              'pourcentage_voix',
            'Parts Indirectes':              'parts_indirectes',
            'Nombre Parts Indirectes':       'nombre_parts_indirectes',
            '% Participation Indirecte':     'pourcentage_participation_indirecte',
            'Droits Vote Indirects':         'droits_vote_indirects',
            'Représentant Légal':            'representant_legal',
            'Representant Legal':            'representant_legal',
            'Date Bénéficiaire':             'date_beneficiaire',
            'Date Beneficiaire':             'date_beneficiaire',
            'Autres Bénéficiaires':          'autres_beneficiaires',
            'Autres Beneficiaires':          'autres_beneficiaires',
            'Certifié Par':                  'certifie_par',
            'Certifie Par':                  'certifie_par',
            'Fonction Certifiant':           'fonction_certifiant',
            'Lien Drive':                    'lien_drive',
            'Statut':                        'statut',
        }

        # Construire le mapping index colonne -> champ modèle
        col_map = {}
        for i, h in enumerate(headers_row):
            if h in MAPPING_COLONNES:
                col_map[i] = MAPPING_COLONNES[h]

        if not col_map:
            flash('Aucune colonne reconnue. Téléchargez le modèle et respectez les en-têtes.', 'danger')
            return redirect(url_for('import_excel'))

        # Détecter la première ligne réelle de données.
        # Le modèle officiel contient exactement :
        #   Ligne 1 : en-têtes des colonnes
        #   Ligne 2 : labels REQUIS / Optionnel
        #   Ligne 3 : exemples illustratifs
        # Un fichier libre n'a qu'une ligne d'en-têtes et les données commencent ligne 2.
        INDICATEURS_METADONNEES = {'requis', 'optionnel', 'obligatoire', 'facultatif', 'example', 'exemple'}

        def premiere_cellule_ligne(row_idx):
            return str(ws.cell(row_idx, 1).value or '').strip().lower()

        premiere_ligne_donnees = 2

        # Si la ligne 2 est une ligne REQUIS/Optionnel -> données à partir de la ligne 4
        # (en sautant aussi la ligne 3 qui est la ligne d'exemples)
        val_ligne2_col1 = premiere_cellule_ligne(2)
        if val_ligne2_col1 in INDICATEURS_METADONNEES:
            premiere_ligne_donnees = 4
        # Cas où la ligne 3 serait aussi REQUIS (structure à 3 lignes de métadonnées)
        elif premiere_cellule_ligne(3) in INDICATEURS_METADONNEES:
            premiere_ligne_donnees = 4

        # Lire les lignes de données
        lignes = []
        erreurs = []
        doublons = []

        for row_idx in range(premiere_ligne_donnees, ws.max_row + 1):
            row_vals = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]

            # Ignorer les lignes complètement vides
            if all(v is None or str(v).strip() == '' for v in row_vals):
                continue

            # Ignorer toute ligne résiduelle de métadonnées
            if premiere_cellule_ligne(row_idx) in INDICATEURS_METADONNEES:
                continue

            record = {}
            for col_idx, champ in col_map.items():
                if col_idx < len(row_vals):
                    val = row_vals[col_idx]
                    if val is None:
                        record[champ] = ''
                    elif isinstance(val, datetime):
                        record[champ] = val.strftime('%Y-%m-%d')
                    else:
                        record[champ] = str(val).strip()
                else:
                    record[champ] = ''

            # Validation minimale
                        
            erreurs_ligne = []
            if not record.get('denomination_sociale'):
                erreurs_ligne.append('Dénomination Sociale manquante')
            if not record.get('nom_beneficiaire'):
                erreurs_ligne.append('Nom Bénéficiaire manquant')

            # Vérifier doublon NINEA si présent
            ninea = record.get('ninea', '').strip()
            if ninea and Declaration.query.filter_by(ninea=ninea).first():
                doublons.append({'ligne': row_idx, 'ninea': ninea,
                                 'entreprise': record.get('denomination_sociale', '')})

            lignes.append({
                'ligne': row_idx,
                'data': record,
                'erreurs': erreurs_ligne,
                'doublon_ninea': bool(ninea and Declaration.query.filter_by(ninea=ninea).first()),
            })

        if not lignes:
            flash('Le fichier ne contient aucune donnée à importer.', 'warning')
            return redirect(url_for('import_excel'))

        # Mode aperçu : retourner un résumé sans importer
        if mode == 'apercu':
            valides = [l for l in lignes if not l['erreurs'] and not l['doublon_ninea']]
            invalides = [l for l in lignes if l['erreurs']]
            return render_template('import_excel.html',
                                   apercu=True,
                                   lignes=lignes,
                                   valides=valides,
                                   invalides=invalides,
                                   doublons=doublons,
                                   total=len(lignes),
                                   colonnes_detectees=list(col_map.values()))

        # Mode import réel
        importees = 0
        ignorees = 0
        user_id = session['user_id']
        ignorer_doublons = request.form.get('ignorer_doublons') == '1'

        for item in lignes:
            if item['erreurs']:
                ignorees += 1
                continue
            if item['doublon_ninea'] and ignorer_doublons:
                ignorees += 1
                continue

            rec = item['data']
            d = Declaration(
                reference=generate_reference(),
                utilisateur_id=user_id,
                statut=rec.get('statut') or 'En attente',
                fichiers='[]',
                **{k: v or None for k, v in rec.items()
                   if k != 'statut' and hasattr(Declaration, k)}
            )
            db.session.add(d)
            importees += 1

        db.session.commit()
        export_to_excel()

        flash(f'{importees} déclaration(s) importée(s) avec succès. {ignorees} ligne(s) ignorée(s).', 'success')
        return redirect(url_for('liste_declarations'))

    except Exception as e:
        flash(f'Erreur lors de la lecture du fichier : {str(e)}', 'danger')
        return redirect(url_for('import_excel'))


@app.route('/import/modele')
@admin_required
def telecharger_modele_import():
    """Génère et télécharge un fichier Excel modèle pré-rempli pour l'import."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Déclarations"

    # Styles
    fill_entete = PatternFill(start_color="1C3557", end_color="1C3557", fill_type="solid")
    fill_requis = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    fill_exemple = PatternFill(start_color="F0F7FF", end_color="F0F7FF", fill_type="solid")
    font_blanc = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    font_requis = Font(bold=True, size=9, color="92400E", name="Calibri")
    font_normal = Font(size=9, name="Calibri")
    bordure = Border(
        left=Side(style='thin', color='CBD5E0'),
        right=Side(style='thin', color='CBD5E0'),
        top=Side(style='thin', color='CBD5E0'),
        bottom=Side(style='thin', color='CBD5E0'),
    )

    # Colonnes du modèle avec métadonnées
    colonnes = [
        # (En-tête, requis, exemple)
        ('Dénomination Sociale',        True,  'TCHILMA BUILDING SARL'),
        ('NINEA',                        True,  '0043385142U2'),
        ('RCCM',                         False, 'SN.DKR.2011.B.2041'),
        ('Forme Juridique',              True,  'SUARL'),
        ('Pays',                         False, 'SÉNÉGAL'),
        ('Région',                       False, 'Dakar'),
        ('Ville',                        False, 'Dakar'),
        ('Adresse',                      False, '76 FASS CASIER, DAKAR'),
        ('Téléphone',                    False, '+221 77 123 45 67'),
        ('Email Entreprise',             False, 'contact@entreprise.sn'),
        ('Cotée en bourse',              False, 'Non'),
        ('Filiale cotée',                False, 'Non'),
        ("Entreprise d'Etat",            False, 'Non'),
        ('Autre entité',                 False, 'Non'),
        ('Nom Maison Mère',              False, ''),
        ('Place Boursière',              False, ''),
        ('Numéro ISIN',                  False, ''),
        ('% Participation Mère',         False, ''),
        ('Nom Bénéficiaire',             True,  'FALL DIOP'),
        ('Prénom Bénéficiaire',          True,  'Ndeye'),
        ('Sexe',                         False, 'Féminin'),
        ('Date Naissance',               False, '09/03/1976'),
        ('Lieu Naissance',               False, 'DAKAR'),
        ('Nationalité',                  False, 'SÉNÉGALAISE'),
        ('Pays Résidence',               False, 'SÉNÉGAL'),
        ('Adresse Personnelle',          False, 'VDN BANDE VERTE LOT 11'),
        ('Adresse Professionnelle',      False, 'FASS CASIER N°76, DAKAR'),
        ('N° CNI / Passeport',           False, '27551976001132'),
        ('PPE',                          False, 'Non'),
        ('Relation PPE',                 False, 'Non'),
        ('Fonction PPE',                 False, ''),
        ('Catégorie PPE',                False, ''),
        ('Date Début Fonction',          False, ''),
        ('Date Fin Fonction',            False, ''),
        ('Parts Directes',               False, 'Oui'),
        ('Nombre Parts Directes',        False, '100'),
        ('% Participation',              False, '100'),
        ('Droits Vote Directs',          False, 'Oui'),
        ('Nombre Voix',                  False, '100'),
        ('% Voix',                       False, '100'),
        ('Parts Indirectes',             False, 'Non'),
        ('Nombre Parts Indirectes',      False, ''),
        ('% Participation Indirecte',    False, ''),
        ('Droits Vote Indirects',        False, 'Non'),
        ('Représentant Légal',           False, 'Oui'),
        ('Date Bénéficiaire',            False, '10/02/2011'),
        ('Autres Bénéficiaires',         False, 'Non'),
        ('Certifié Par',                 True,  'FALL DIOP Ndeye'),
        ('Fonction Certifiant',          True,  'GERANTE'),
        ('Lien Drive',                   False, ''),
        ('Statut',                       False, 'En attente'),
    ]

    # Ligne 1 : en-têtes
    for col, (titre, requis, _) in enumerate(colonnes, 1):
        cell = ws.cell(row=1, column=col, value=titre)
        cell.fill = fill_entete
        cell.font = font_blanc
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = bordure
        if requis:
            # Marquer les colonnes requises avec un fond distinct
            cell.fill = PatternFill(start_color="0F2340", end_color="0F2340", fill_type="solid")

    ws.row_dimensions[1].height = 35

    # Ligne 2 : indication requis/optionnel
    for col, (_, requis, _) in enumerate(colonnes, 1):
        cell = ws.cell(row=2, column=col, value='REQUIS' if requis else 'Optionnel')
        cell.fill = fill_requis if requis else PatternFill(start_color="F0FFF4", end_color="F0FFF4", fill_type="solid")
        cell.font = font_requis if requis else Font(size=8, color="276749", name="Calibri", italic=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bordure

    # Ligne 3 : exemple de données
    for col, (_, _, exemple) in enumerate(colonnes, 1):
        cell = ws.cell(row=3, column=col, value=exemple)
        cell.fill = fill_exemple
        cell.font = Font(size=9, color="2A4D7A", name="Calibri", italic=True)
        cell.alignment = Alignment(vertical='center')
        cell.border = bordure

    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18

    # Lignes vides pour saisie (4 à 103 pour 100 lignes)
    fill_alt = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    for row in range(4, 104):
        for col in range(1, len(colonnes) + 1):
            cell = ws.cell(row=row, column=col, value='')
            cell.border = bordure
            cell.font = font_normal
            cell.alignment = Alignment(vertical='center')
            if row % 2 == 0:
                cell.fill = fill_alt

    # Largeur des colonnes
    largeurs = [28, 18, 22, 16, 12, 12, 12, 30, 18, 24, 14, 14, 14, 14,
                22, 18, 16, 16, 22, 22, 12, 14, 16, 18, 16, 28, 28, 20,
                10, 12, 18, 20, 16, 16, 14, 18, 14, 16, 14, 10, 14, 18,
                18, 16, 16, 16, 18, 22, 22, 28, 14]
    for col, larg in enumerate(largeurs, 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = larg

    # Figer les deux premières lignes et la première colonne
    ws.freeze_panes = 'B4'

    # Onglet instructions
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ("INSTRUCTIONS D'UTILISATION DU MODÈLE", True),
        ("", False),
        ("1. Ne pas modifier les en-têtes de la ligne 1 (noms des colonnes).", False),
        ("2. La ligne 2 indique les champs REQUIS (obligatoires) et Optionnels.", False),
        ("3. La ligne 3 contient des exemples à titre indicatif — supprimez-la avant d'importer.", False),
        ("4. Saisir les données à partir de la ligne 4.", False),
        ("5. Pour les champs Oui/Non, utiliser exactement : Oui, Non, Non renseigné", False),
        ("6. Dates au format DD/MM/YYYY ou YYYY-MM-DD.", False),
        ("7. Le champ NINEA doit être unique dans toute la base.", False),
        ("8. Le champ Statut accepte : En attente, Validée, Rejetée", False),
        ("9. Ne pas modifier la structure de ce fichier (ne pas ajouter ni supprimer de colonnes).", False),
        ("", False),
        ("VALEURS ACCEPTÉES", True),
        ("Sexe : Masculin / Féminin / Non renseigné", False),
        ("Forme Juridique : SA, SARL, SUARL, SNC, SCS, GIE, Coopérative, Association, ONG, Autre", False),
        ("Région : Dakar, Thiès, Diourbel, Fatick, Kaolack, Kaffrine, Louga, Saint-Louis, Matam,", False),
        ("         Tambacounda, Kédougou, Kolda, Sédhiou, Ziguinchor, Autre", False),
        ("Catégorie PPE : PPE nationale / PPE étrangère / PPE organisation internationale", False),
    ]
    ws2.column_dimensions['A'].width = 80
    fill_titre = PatternFill(start_color="1C3557", end_color="1C3557", fill_type="solid")
    for i, (texte, titre) in enumerate(instructions, 1):
        cell = ws2.cell(row=i, column=1, value=texte)
        if titre:
            cell.font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
            cell.fill = fill_titre
        else:
            cell.font = Font(size=10, name="Calibri", color="1A202C")
        cell.alignment = Alignment(vertical='center')
        ws2.row_dimensions[i].height = 18

    exports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')
    os.makedirs(exports_dir, exist_ok=True)
    chemin = os.path.join(exports_dir, 'modele_import_declarations.xlsx')
    wb.save(chemin)
    return send_file(chemin, as_attachment=True, download_name='modele_import_declarations.xlsx')


@app.route('/api/stats')
@login_required
def api_stats():
    user = User.query.get(session['user_id'])
    uid = None if user.role == 'admin' else user.id
    return jsonify(get_stats(user_id=uid))


@app.route('/api/validate/ninea', methods=['POST'])
@login_required
def validate_ninea():
    ninea = request.json.get('ninea', '')
    existing = Declaration.query.filter_by(ninea=ninea).first()
    if existing:
        return jsonify({'valid': False, 'message': f'Ce NINEA est déjà enregistré (Réf. {existing.reference})'})
    return jsonify({'valid': True})


def export_to_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Déclarations"

    header_fill = PatternFill(start_color="1C3557", end_color="1C3557", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    border = Border(
        left=Side(style='thin', color='CBD5E0'),
        right=Side(style='thin', color='CBD5E0'),
        top=Side(style='thin', color='CBD5E0'),
        bottom=Side(style='thin', color='CBD5E0')
    )

    headers = ['ID', 'Référence', 'Date', 'Utilisateur', 'Email', 'Dénomination Sociale',
               'NINEA', 'RCCM', 'Forme Juridique', 'Pays', 'Région', 'Adresse',
               'Nom Bénéficiaire', 'Prénom Bénéficiaire', 'Nationalité',
               'PPE', 'Participation %', 'Statut', 'Lien Drive']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 30

    declarations = Declaration.query.all()
    for row, d in enumerate(declarations, 2):
        u = User.query.get(d.utilisateur_id)
        valeurs = [
            d.id, d.reference,
            d.date_saisie.strftime('%d/%m/%Y %H:%M') if d.date_saisie else '',
            f"{u.prenom} {u.nom}" if u else '', u.email if u else '',
            d.denomination_sociale, d.ninea, d.rccm, d.forme_juridique, d.pays, d.region,
            d.adresse, d.nom_beneficiaire, d.prenom_beneficiaire, d.nationalite,
            d.ppe, d.pourcentage_participation, d.statut, d.lien_drive or ''
        ]
        bg = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid") if row % 2 == 0 else None
        for col, val in enumerate(valeurs, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.font = Font(size=9, name="Calibri")
            cell.alignment = Alignment(vertical='center')
            if bg:
                cell.fill = bg

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    exports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')
    os.makedirs(exports_dir, exist_ok=True)
    path = os.path.join(exports_dir, 'registre_beneficiaires_effectifs.xlsx')
    wb.save(path)
    return path


def init_db():
    with app.app_context():
        db.create_all()
        if not User.query.filter_by(role='admin').first():
            hashed = bcrypt.generate_password_hash('changez_ce_mot_de_passe').decode('utf-8')
            admin_user = User(nom='Administrateur', prenom='Système',
                              email='admin@rne.justice.sn', password=hashed, role='admin')
            db.session.add(admin_user)
            db.session.commit()


# Filtre Jinja personnalisé
@app.template_filter("from_json")
def from_json_filter(value):
    if not value:
        return []
    try:
        return json.loads(value)
    except Exception:
        return []


def initialiser_base():
    """Crée les tables et le compte admin si nécessaires."""
    with app.app_context():
        db.create_all()
        if not User.query.filter_by(role='admin').first():
            _hashed = bcrypt.generate_password_hash('changez_ce_mot_de_passe').decode('utf-8')
            _admin = User(
                nom='Administrateur', prenom='Système',
                email='admin@rne.justice.sn',
                password=_hashed, role='admin'
            )
            db.session.add(_admin)
            db.session.commit()
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        _exports = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')
        os.makedirs(_exports, exist_ok=True)


# Exécuté au chargement du module (gunicorn, python app.py, tests)
initialiser_base()


@app.before_request
def verifier_base():
    """Filet de sécurité : recrée les tables si elles ont disparu."""
    try:
        User.query.limit(1).all()
    except Exception:
        db.create_all()


if __name__ == "__main__":
    app.run(debug=False, host='0.0.0.0', port=5000)
